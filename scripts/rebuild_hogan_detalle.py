import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import json
import os
import sys
import unicodedata
import re
import openpyxl

BASE_DIR = "/Users/nicolassantos/Desktop/Proyectos/RV_ Reportes  Hogan Final "
OUT_FULL = os.path.join(os.path.dirname(__file__), "hogan_detalle_full.json")
OUT_KEYED = os.path.join(os.path.dirname(__file__), "hogan_detalle_keyed.json")

FILES = [
    "Direccion_Auditoria_Fichas_v2.xlsx",
    "Direccion_Financiera_v2.xlsx",
    "Direccion_Gestion_Humana_v2.xlsx",
    "Direccion_Juridica_v2.xlsx",
    "Direccion_Planeacion_Estrategica_v2.xlsx",
    "Direccion_Riesgo_v2.xlsx",
    "Direccion_Tecnologia_v2.xlsx",
    "Presidencia_v2.xlsx",
    "Unidad_Mercado_Corporativo_v2.xlsx",
    "Unidad_Mercado_Masivo_v2.xlsx",
]

SKIP_SHEETS = {
    "Buscador", "Índice General", "Análisis de Grupo",
    "Dic_LE", "Dic_LG", "Data_LE", "Data_LG", "Lista_Personas",
}

FIELD_MAP = {
    "Dirección Corporativa": "direccion_corporativa",
    "Dirección Área": "direccion_area",
    "Gerencia": "gerencia",
    "Región": "region",
    "Jefe": "jefe",
    "Nivel de Contribución": "nivel_contribucion",
}

STOP = {'DE', 'LA', 'DEL', 'LOS', 'LAS', 'Y', 'SAN', 'SANTA'}


def norm_words(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    words = [w for w in re.split(r'\s+', s) if w and w not in STOP]
    words.sort()
    return ' '.join(words)


def get_rows(ws):
    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append((idx, list(row)))
    return rows


def first_nonempty_cell(row):
    for v in row:
        if v is not None and str(v).strip() != "":
            return v
    return None


def row_values_nonempty(row):
    return [v for v in row if v is not None and str(v).strip() != ""]


def parse_sheet(ws, source_file, source_sheet):
    rows = get_rows(ws)

    record = {
        "nombre": None,
        "cargo": None,
        "direccion_corporativa": None,
        "direccion_area": None,
        "gerencia": None,
        "region": None,
        "jefe": None,
        "nivel_contribucion": None,
        "competencias": [],
        "fortalezas": None,
        "oportunidades": None,
        "recomendaciones": None,
        "recomendaciones_nota": None,
        "source_file": source_file,
        "source_sheet": source_sheet,
    }

    competencias_header_idx = None
    for idx, row in rows:
        v = first_nonempty_cell(row)
        if v is not None and str(v).strip() == "COMPETENCIAS":
            competencias_header_idx = idx
            break

    single_value_rows = []
    two_col_start_idx = None
    for idx, row in rows:
        if competencias_header_idx is not None and idx >= competencias_header_idx:
            break
        if idx == 1:
            continue
        vals = row_values_nonempty(row)
        if len(vals) == 1:
            single_value_rows.append((idx, vals[0]))
        elif len(vals) == 2 and str(vals[0]).strip() in FIELD_MAP:
            if two_col_start_idx is None:
                two_col_start_idx = idx
        if two_col_start_idx is not None:
            if len(vals) == 1:
                break

    name_cargo_candidates = [(idx, v) for idx, v in single_value_rows if two_col_start_idx is None or idx < two_col_start_idx]
    if len(name_cargo_candidates) >= 1:
        record["nombre"] = str(name_cargo_candidates[0][1]).strip()
    if len(name_cargo_candidates) >= 2:
        record["cargo"] = str(name_cargo_candidates[1][1]).strip()

    for idx, row in rows:
        if competencias_header_idx is not None and idx >= competencias_header_idx:
            break
        vals = row_values_nonempty(row)
        if len(vals) == 2 and str(vals[0]).strip() in FIELD_MAP:
            key = FIELD_MAP[str(vals[0]).strip()]
            record[key] = str(vals[1]).strip() if vals[1] is not None else None

    if competencias_header_idx is not None:
        header_idx = None
        for idx, row in rows:
            if idx <= competencias_header_idx:
                continue
            vals = row_values_nonempty(row)
            if not vals:
                continue
            if len(vals) >= 2 and str(vals[0]).strip() == "Competencia" and str(vals[1]).strip() == "Puntaje":
                header_idx = idx
            break
        if header_idx is not None:
            for idx, row in rows:
                if idx <= header_idx:
                    continue
                vals = row_values_nonempty(row)
                if not vals:
                    break
                if len(vals) < 3:
                    break
                nombre_c, puntaje_c, desc_c = vals[0], vals[1], vals[2]
                try:
                    puntaje_int = int(puntaje_c)
                except (ValueError, TypeError):
                    puntaje_int = puntaje_c
                # In the RV_ files the competency cell is "Nombre\nDefinición".
                # Keep only the first line as the name (schema parity with the
                # existing hogan_detalle and with how the tableros render it).
                nombre_clean = str(nombre_c).strip().split("\n")[0].strip()
                record["competencias"].append({
                    "nombre": nombre_clean,
                    "puntaje": puntaje_int,
                    "descripcion": str(desc_c).strip() if desc_c is not None else None,
                })

    fortalezas_header_idx = None
    for idx, row in rows:
        v = first_nonempty_cell(row)
        if v is not None and str(v).strip() == "FORTALEZAS":
            fortalezas_header_idx = idx
            break
    if fortalezas_header_idx is not None:
        for idx, row in rows:
            if idx <= fortalezas_header_idx:
                continue
            vals = row_values_nonempty(row)
            if not vals:
                continue
            record["fortalezas"] = str(vals[0]).strip()
            break

    oportunidades_header_idx = None
    for idx, row in rows:
        v = first_nonempty_cell(row)
        if v is not None and str(v).strip() == "OPORTUNIDADES":
            oportunidades_header_idx = idx
            break
    if oportunidades_header_idx is not None:
        for idx, row in rows:
            if idx <= oportunidades_header_idx:
                continue
            vals = row_values_nonempty(row)
            if not vals:
                continue
            record["oportunidades"] = str(vals[0]).strip()
            break

    reco_header_idx = None
    for idx, row in rows:
        v = first_nonempty_cell(row)
        if v is not None and str(v).strip() == "RECOMENDACIONES DE DESARROLLO Y EXPERIENCIAS":
            reco_header_idx = idx
            break

    if reco_header_idx is not None:
        next_idx = None
        next_vals = None
        for idx, row in rows:
            if idx <= reco_header_idx:
                continue
            vals = row_values_nonempty(row)
            if not vals:
                continue
            next_idx = idx
            next_vals = vals
            break

        is_table = (
            next_vals is not None
            and len(next_vals) >= 2
            and str(next_vals[0]).strip() == "Competencia"
            and str(next_vals[1]).strip() == "Experiencias clave / opciones de desarrollo"
        )

        if is_table:
            recomendaciones = []
            for idx, row in rows:
                if idx <= next_idx:
                    continue
                vals = row_values_nonempty(row)
                if not vals:
                    break
                if len(vals) < 2:
                    break
                comp, exp_text = vals[0], vals[1]
                bullets = [b.strip() for b in str(exp_text).split("•") if b.strip()]
                recomendaciones.append({
                    "competencia": str(comp).strip(),
                    "experiencias": bullets,
                })
            record["recomendaciones"] = recomendaciones
        else:
            texts = []
            for idx, row in rows:
                if idx <= reco_header_idx:
                    continue
                vals = row_values_nonempty(row)
                if not vals:
                    continue
                texts.append(str(vals[0]).strip())
            record["recomendaciones"] = None
            record["recomendaciones_nota"] = "\n".join(texts) if texts else None

    return record


def main():
    all_records = []
    total_sheets = 0
    fully_parsed = 0
    edge_case_list = []
    failed_list = []

    for fname in FILES:
        path = os.path.join(BASE_DIR, fname)
        if not os.path.exists(path):
            print(f"WARNING: file not found: {path}")
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("F_"):
                continue
            if sheet_name in SKIP_SHEETS:
                continue
            total_sheets += 1
            try:
                ws = wb[sheet_name]
                record = parse_sheet(ws, fname, sheet_name)
                all_records.append(record)
                if record["nombre"] and record["competencias"]:
                    fully_parsed += 1
                if record["recomendaciones"] is None and record["recomendaciones_nota"]:
                    edge_case_list.append({
                        "source_file": fname,
                        "source_sheet": sheet_name,
                        "nombre": record["nombre"],
                    })
            except Exception as e:
                failed_list.append({
                    "source_file": fname,
                    "source_sheet": sheet_name,
                    "error": f"{type(e).__name__}: {e}",
                })
        wb.close()

    with open(OUT_FULL, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    keyed = {}
    dupes = []
    for r in all_records:
        key = norm_words(r["nombre"])
        if not key:
            continue
        if key in keyed:
            dupes.append((key, r["source_file"], r["source_sheet"]))
        keyed[key] = r

    with open(OUT_KEYED, "w", encoding="utf-8") as f:
        json.dump(keyed, f, ensure_ascii=False)

    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total F_ sheets processed: {total_sheets}")
    print(f"Total records written: {len(all_records)}")
    print(f"Fully parsed (nombre + competencias non-empty): {fully_parsed}")
    print(f"Unique normalized keys: {len(keyed)}")
    print(f"Duplicate keys: {len(dupes)}")
    for d in dupes:
        print(f"  - {d}")
    print(f"Edge-case (recomendaciones_nota, no table): {len(edge_case_list)}")
    for e in edge_case_list:
        print(f"  - {e['source_file']} / {e['source_sheet']} / {e['nombre']}")
    print(f"Failed to parse: {len(failed_list)}")
    for e in failed_list:
        print(f"  - {e['source_file']} / {e['source_sheet']}: {e['error']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
