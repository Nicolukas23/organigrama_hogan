"""
sync_engine.py — Motor de diff + sync Supabase
Lee Excels, compara con data actual en Supabase, detecta cambios y aplica solo los diffs.
"""
import os
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from supabase import create_client

# ── Config ──────────────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
BATCH_SIZE = 500


def get_client():
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ── Excel helpers ───────────────────────────────────────
_wb_cache: dict[str, openpyxl.Workbook] = {}


def _read_wb(path: str) -> openpyxl.Workbook:
    if path not in _wb_cache:
        _wb_cache[path] = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return _wb_cache[path]


def read_sheet(path: str, sheet_name: str) -> list[list]:
    wb = _read_wb(path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _find_header_row(data: list[list], needle: str) -> int:
    for i, row in enumerate(data[:5]):
        if row and any(
            needle in str(c).lower() for c in row if c is not None
        ):
            return i
    return -1


def _headers(data: list[list], idx: int) -> list[str]:
    return [str(c or "").lower().strip() for c in data[idx]]


def _col(headers: list[str], *needles: str) -> int:
    for i, h in enumerate(headers):
        for n in needles:
            if n in h:
                return i
    return -1


def _val(row: list, idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    return v if v is not None else None


# ── Extract: Ficha de Talento ──────────────────────────
def extract_personas(ficha1: str, planta: str) -> dict[str, dict]:
    info: dict[str, dict] = {}

    # Fuente secundaria: Planta
    planta_map: dict[str, dict] = {}
    plant_data = read_sheet(planta, "Planta Claro")
    if plant_data:
        ph = _find_header_row(plant_data, "expediente")
        if ph >= 0:
            phd = _headers(plant_data, ph)
            pc = {
                "expediente": _col(phd, "expediente"),
                "nombre": _col(phd, "apellidos y nombres"),
                "cargo": _col(phd, "cargo"),
                "empresa": _col(phd, "compañia"),
                "region": _col(phd, "region gv"),
                "ciudades": _col(phd, "ciudad"),
                "estado": _col(phd, "estado"),
            }
            for row in plant_data[ph + 1 :]:
                exp_val = _val(row, pc["expediente"])
                if not exp_val:
                    continue
                exp = str(exp_val).strip()
                if not exp or exp == "undefined":
                    continue
                planta_map[exp] = {
                    "nombre": _val(row, pc["nombre"]) or "",
                    "cargo": _val(row, pc["cargo"]) or "",
                    "empresa": _val(row, pc["empresa"]) or "",
                    "region": _val(row, pc["region"]) or "",
                    "ciudades": _val(row, pc["ciudades"]) or "",
                    "estado": _val(row, pc["estado"]) or "",
                }

    # Fuente principal: Ficha 1
    data = read_sheet(ficha1, "INFORMACION")
    if not data:
        return info
    hr = _find_header_row(data, "cedula")
    if hr < 0:
        return info
    h = _headers(data, hr)
    c = {
        "expediente": _col(h, "cedula"),
        "nombre": _col(h, "nombre"),
        "direccion_comite": _col(h, "direccion comite"),
        "area": _col(h, "direccion area"),
        "gerencia": _col(h, "gerencia"),
        "fecha_ingreso": _col(h, "fecha de ingreso"),
        "antiguedad": _col(h, "antig"),
        "jefe": _col(h, "jefe"),
    }
    for row in data[hr + 1 :]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp or exp == "undefined":
            continue
        pl = planta_map.get(exp, {})
        info[exp] = {
            "expediente": exp,
            "nombre": _val(row, c["nombre"]) or pl.get("nombre", ""),
            "cargo": pl.get("cargo", ""),
            "empresa": pl.get("empresa", ""),
            "region": pl.get("region", ""),
            "ciudades": pl.get("ciudades", ""),
            "estado": pl.get("estado", ""),
            "direccion_comite": _val(row, c["direccion_comite"]) or "",
            "area": _val(row, c["area"]) or "",
            "gerencia": _val(row, c["gerencia"]) or "",
            "fecha_ingreso": _val(row, c["fecha_ingreso"]) or "",
            "antiguedad": _val(row, c["antiguedad"]) or "",
            "jefe": _val(row, c["jefe"]) or "",
        }
    return info


def extract_formacion(ficha1: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    data = read_sheet(ficha1, "Formación&Experiencia")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr <= 0:
        return out
    h = _headers(data, hr - 1)
    c = {
        "expediente": _col(h, "expediente"),
        "educacion_formal": _col(h, "educación formal"),
        "educacion_complementaria": _col(h, "educación complementaria"),
        "experiencia_claro": _col(h, "experiencia claro"),
        "experiencia_otros": _col(h, "experiencia otros"),
    }
    for row in data[hr:]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        out[exp] = {
            "expediente": exp,
            "educacion_formal": _val(row, c["educacion_formal"]) or "",
            "educacion_complementaria": _val(row, c["educacion_complementaria"]) or "",
            "experiencia_claro": _val(row, c["experiencia_claro"]) or "",
            "experiencia_otros": _val(row, c["experiencia_otros"]) or "",
        }
    return out


def extract_desempeno(ficha1: str, ficha2: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    years = [(ficha1, 2024), (ficha1, 2025), (ficha2, 2026)]
    for fpath, year in years:
        data = read_sheet(fpath, "Desempeño")
        if not data:
            continue
        hr = _find_header_row(data, "expediente")
        if hr < 0:
            continue
        h = _headers(data, hr)
        exp_idx = _col(h, "expediente")
        puntaje_idx = _col(h, "puntaje")
        desempeno_idx = _col(h, "desempeño", "desempeno")
        for row in data[hr + 1 :]:
            exp_val = _val(row, exp_idx)
            if not exp_val:
                continue
            exp = str(exp_val).strip()
            if not exp:
                continue
            if exp not in out:
                out[exp] = {"expediente": exp, "y2024": None, "y2025": None, "y2026": None}
            v = _val(row, puntaje_idx) or _val(row, desempeno_idx)
            out[exp][f"y{year}"] = v
    return out


def extract_talentos(ficha2: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    data = read_sheet(ficha2, "Talentos")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    c = {
        "expediente": _col(h, "expediente"),
        "talento": _col(h, "talento"),
        "soy_dueno": _col(h, "soy dueño"),
        "soy_lider": _col(h, "soy líder", "soy lider"),
        "soy_digital": _col(h, "soy digital"),
    }
    for row in data[hr + 1 :]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        out[exp] = {
            "expediente": exp,
            "talento": _val(row, c["talento"]) or "",
            "soy_dueno": _val(row, c["soy_dueno"]),
            "soy_lider": _val(row, c["soy_lider"]),
            "soy_digital": _val(row, c["soy_digital"]),
        }
    return out


def extract_objetivos(ficha2: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    data = read_sheet(ficha2, "Objetivo Desarrollo")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    c = {
        "expediente": _col(h, "expediente"),
        "obj": _col(h, "objetivo"),
        "talento": _col(h, "talento"),
    }
    for row in data[hr + 1 :]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        out[exp] = {
            "expediente": exp,
            "obj": _val(row, c["obj"]) or "",
            "talento": _val(row, c["talento"]) or "",
        }
    return out


def extract_360(ficha2: str) -> list[dict]:
    out: list[dict] = []
    data = read_sheet(ficha2, "360")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    c = {
        "expediente": _col(h, "expediente"),
        "competencia": _col(h, "competencia", "comportamiento"),
        "lider": _col(h, "líder", "lider"),
        "propio": _col(h, "propio"),
        "reporte": _col(h, "reporte"),
        "par": _col(h, "par"),
    }
    for row in data[hr + 1 :]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        out.append({
            "expediente": exp,
            "competencia": _val(row, c["competencia"]) or "",
            "lider": _val(row, c["lider"]),
            "propio": _val(row, c["propio"]),
            "reporte": _val(row, c["reporte"]),
            "par": _val(row, c["par"]),
        })
    return out


def extract_clima(ficha2: str) -> list[dict]:
    out: list[dict] = []
    data = read_sheet(ficha2, "Clima")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    c = {
        "expediente": _col(h, "expediente"),
        "dimension": _col(h, "dimensión", "dimension"),
        "pct": _col(h, "porcentaje", "pct"),
    }
    for row in data[hr + 1 :]:
        exp_val = _val(row, c["expediente"])
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        out.append({
            "expediente": exp,
            "dimension": _val(row, c["dimension"]) or "",
            "pct": _val(row, c["pct"]),
        })
    return out


# ── Extract: Nine Box ──────────────────────────────────
def extract_sucesores(ninebox: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    data = read_sheet(ninebox, "observaciones")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    exp_idx = _col(h, "expediente")
    sucesor_idx = 25
    tiempo_idx = 26
    for row in data[hr + 1 :]:
        exp_val = _val(row, exp_idx)
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        sucesor = _val(row, sucesor_idx)
        tiempo = _val(row, tiempo_idx)
        if sucesor or tiempo:
            out[exp] = {"expediente": exp, "sucesor": sucesor or "", "tiempo": tiempo or ""}
    return out


def extract_ninebox(ninebox: str) -> list[dict]:
    """Extrae las filas completas del Nine Box para la tabla ninebox."""
    out: list[dict] = []
    data = read_sheet(ninebox, "observaciones")
    if not data:
        return out
    hr = _find_header_row(data, "expediente")
    if hr < 0:
        return out
    h = _headers(data, hr)
    for row in data[hr + 1 :]:
        exp_val = _val(row, _col(h, "expediente"))
        if not exp_val:
            continue
        exp = str(exp_val).strip()
        if not exp:
            continue
        rec: dict[str, Any] = {"expediente": exp}
        for i, hdr in enumerate(h):
            if i >= len(row):
                continue
            val = row[i]
            if val is not None:
                rec[hdr] = val
        out.append(rec)
    return out


# ── Diff engine ────────────────────────────────────────
class DiffResult:
    """Resultado del diff para una tabla."""

    def __init__(self, table: str):
        self.table = table
        self.added: list[dict] = []
        self.removed: list[dict] = []
        self.modified: list[dict] = []  # {row, campo, antes, despues}

    @property
    def total_changes(self) -> int:
        return len(self.added) + len(self.removed) + len(self.modified)

    @property
    def has_changes(self) -> bool:
        return self.total_changes > 0

    def summary(self) -> dict:
        return {
            "tabla": self.table,
            "agregados": len(self.added),
            "eliminados": len(self.removed),
            "modificados": len(self.modified),
            "total": self.total_changes,
        }


def compute_diff(
    old_data: dict[str, dict] | list[dict],
    new_data: dict[str, dict] | list[dict],
    table: str,
    key: str = "expediente",
) -> DiffResult:
    """Compara old vs new y retorna un DiffResult."""
    diff = DiffResult(table)

    # Normalizar a dict keyed by expediente
    def _to_dict(data):
        if isinstance(data, list):
            return {str(d.get(key, "")): d for d in data if d.get(key)}
        return data

    old_d = _to_dict(old_data)
    new_d = _to_dict(new_data)

    # Detectar agregados y modificados
    for k, new_row in new_d.items():
        if k not in old_d:
            diff.added.append(new_row)
        else:
            old_row = old_d[k]
            changes = []
            for field in new_row:
                if field == key or field == "updated_at":
                    continue
                old_val = old_row.get(field)
                new_val = new_row.get(field)
                if str(old_val) != str(new_val):
                    changes.append({"campo": field, "antes": old_val, "despues": new_val})
            if changes:
                diff.modified.append({"expediente": k, "nombre": new_row.get("nombre", ""), "cambios": changes})

    # Detectar eliminados
    for k in old_d:
        if k not in new_d:
            diff.removed.append(old_d[k])

    return diff


def compute_diff_flat(
    old_data: list[dict],
    new_data: list[dict],
    table: str,
) -> DiffResult:
    """Para tablas planas (360, clima): compara por contenido completo."""
    diff = DiffResult(table)

    def _fingerprint(row: dict) -> str:
        return json.dumps({k: v for k, v in sorted(row.items()) if k != "updated_at"}, default=str)

    old_fps = {_fingerprint(r): r for r in old_data}
    new_fps = {_fingerprint(r): r for r in new_data}

    for fp, row in new_fps.items():
        if fp not in old_fps:
            diff.added.append(row)

    for fp, row in old_fps.items():
        if fp not in new_fps:
            diff.removed.append(row)

    return diff


# ── Supabase fetch ─────────────────────────────────────
def fetch_current(client, table: str, key: str = "expediente") -> dict[str, dict] | list[dict]:
    """Trae toda la data actual de una tabla Supabase."""
    all_rows: list[dict] = []
    offset = 0
    while True:
        resp = client.table(table).select("*").range(offset, offset + BATCH_SIZE - 1).execute()
        rows = resp.data or []
        all_rows.extend(rows)
        if len(rows) < BATCH_SIZE:
            break
        offset += BATCH_SIZE
    return {str(r.get(key, "")): r for r in all_rows if r.get(key)}


def fetch_current_flat(client, table: str) -> list[dict]:
    """Trae toda la data de tablas planas (360, clima)."""
    all_rows: list[dict] = []
    offset = 0
    while True:
        resp = client.table(table).select("*").range(offset, offset + BATCH_SIZE - 1).execute()
        rows = resp.data or []
        all_rows.extend(rows)
        if len(rows) < BATCH_SIZE:
            break
        offset += BATCH_SIZE
    return all_rows


# ── Supabase sync ──────────────────────────────────────
def _clean_row(row: dict) -> dict:
    """Limpia strings vacíos a None para evitar errores de tipos."""
    return {k: (None if v == "" else v) for k, v in row.items()}


def upsert_rows(client, table: str, rows: list[dict], replace_all: bool = False):
    """Upsert filas a Supabase en chunks."""
    if not rows:
        return

    now = datetime.now(timezone.utc).isoformat()

    if replace_all:
        # Eliminar todo y re-insertar
        client.table(table).delete().neq("expediente", "__none__").execute()
        for i in range(0, len(rows), BATCH_SIZE):
            chunk = [_clean_row({**r, "updated_at": now}) for r in rows[i : i + BATCH_SIZE]]
            client.table(table).insert(chunk).execute()
    else:
        for i in range(0, len(rows), BATCH_SIZE):
            chunk = [_clean_row({**r, "updated_at": now}) for r in rows[i : i + BATCH_SIZE]]
            client.table(table).upsert(chunk, on_conflict="expediente").execute()


# ── Full sync orchestrator ─────────────────────────────
class SyncReport:
    """Reporte completo del sync."""

    def __init__(self):
        self.diffs: list[DiffResult] = []
        self.start_time = datetime.now()
        self.end_time: datetime | None = None
        self.errors: list[str] = []
        self.files_used: list[str] = []

    @property
    def total_changes(self) -> int:
        return sum(d.total_changes for d in self.diffs)

    @property
    def has_changes(self) -> bool:
        return any(d.has_changes for d in self.diffs)

    def summary(self) -> dict:
        return {
            "fecha": self.start_time.isoformat(),
            "archivos": self.files_used,
            "tablas": [d.summary() for d in self.diffs if d.has_changes],
            "total_cambios": self.total_changes,
            "errores": self.errors,
        }


def run_full_sync(
    ficha1_path: str,
    ficha2_path: str,
    ninebox_path: str,
    planta_path: str | None = None,
    dry_run: bool = False,
) -> SyncReport:
    """
    Ejecuta el sync completo:
    1. Extrae data de Excels
    2. Compara con Supabase
    3. Retorna reporte de diff
    4. Si dry_run=False, aplica los cambios
    """
    report = SyncReport()
    report.files_used = [Path(f).name for f in [ficha1_path, ficha2_path, ninebox_path] if f]

    if not planta_path:
        # Buscar Planta_de_Personal en el directorio padre
        parent = Path(ficha1_path).parent
        candidates = list(parent.glob("Planta_de_Personal*.xlsx"))
        planta_path = str(candidates[0]) if candidates else ficha1_path

    print("📥 Extrayendo data de Excels...")
    personas = extract_personas(ficha1_path, planta_path)
    formacion = extract_formacion(ficha1_path)
    desempeno = extract_desempeno(ficha1_path, ficha2_path)
    talentos = extract_talentos(ficha2_path)
    objetivos = extract_objetivos(ficha2_path)
    tres60 = extract_360(ficha2_path)
    clima = extract_clima(ficha2_path)
    sucesores = extract_sucesores(ninebox_path)

    print(f"  personas={len(personas)}, formacion={len(formacion)}, desempeno={len(desempeno)}, "
          f"talentos={len(talentos)}, objetivos={len(objetivos)}, 360={len(tres60)}, "
          f"clima={len(clima)}, sucesores={len(sucesores)}")

    if dry_run:
        print("\n🔍 Modo dry run — solo analizando cambios...")
        # En dry run, no conectamos a Supabase
        # Simulamos diffs vacíos (la app hará fetch por separado)
        return report

    print("\n🔗 Conectando a Supabase...")
    client = get_client()

    # Definir tablas y sus extractores
    tables = [
        ("personas", personas, False),
        ("formacion", formacion, False),
        ("desempeno", desempeno, False),
        ("talentos", talentos, False),
        ("objetivos", objetivos, False),
        ("sucesores", sucesores, False),
    ]

    for table, new_data, replace_all in tables:
        print(f"\n📊 {table}:")
        try:
            old = fetch_current(client, table)
            diff = compute_diff(old, new_data, table)
            report.diffs.append(diff)
            print(f"  +{len(diff.added)} agregados, -{len(diff.removed)} eliminados, ↻{len(diff.modified)} modificados")
            if not dry_run and diff.has_changes:
                rows = list(new_data.values()) if isinstance(new_data, dict) else new_data
                upsert_rows(client, table, rows, replace_all=replace_all)
                print(f"  ✅ Sync OK ({len(rows)} filas)")
        except Exception as e:
            report.errors.append(f"{table}: {str(e)}")
            print(f"  ❌ Error: {e}")

    # Tablas planas (replace all)
    for table, new_data in [("evaluacion_360", tres60), ("clima", clima)]:
        print(f"\n📊 {table}:")
        try:
            old = fetch_current_flat(client, table)
            diff = compute_diff_flat(old, new_data, table)
            report.diffs.append(diff)
            print(f"  +{len(diff.added)} nuevos, -{len(diff.removed)} eliminados")
            if not dry_run:
                upsert_rows(client, table, new_data, replace_all=True)
                print(f"  ✅ Sync OK ({len(new_data)} filas)")
        except Exception as e:
            report.errors.append(f"{table}: {str(e)}")
            print(f"  ❌ Error: {e}")

    report.end_time = datetime.now()
    print(f"\n{'='*50}")
    print(f"✅ SYNC COMPLETADO — {report.total_changes} cambios detectados")
    if report.errors:
        print(f"⚠️  {len(report.errors)} errores")

    return report
